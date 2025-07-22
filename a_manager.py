import random
import json
import datetime
import openpyxl
from tkinter import ttk, messagebox, Toplevel, Text, Scrollbar, Label, Button
from tkinter.font import Font
import uuid
import math
import textwrap

Adventure_Feature_Enabled = True
MAX_ADVENTURES = 5 #adventures to keep in the log
MAX_ENCOUNTERS = 100 #how many enemies can be encountered in one adventure after skips
STAGES_PER_DIFFICULTY = 3
STAGES_PER_GEAR_LEVEL = 10
DIFFICULTY_INCREASE_PER_STAGE = 0.01
DIFFICULTY_RANGE = 30
STAGES_PER_BOSS = 100
HIGHEST_STAGE = 50000 #how many enemies can be encountered in one adventure including skips
XP_PER_KILL = 3
NORMAL_HIT_VARIATION = 0.05 #how much player attack damage will vary
TRACKER_FILE_PATH = "./Assets/adventure_tracker.json"
LOG_FILE_PATH = "./Assets/adventure_log.txt"
ADVENTURER_FILE_PATH = "./Assets/adventurer.json"
LEADERBOARD_FILE_PATH = "./Assets/leaderboard.json"
CONTENT_FILE_PATH = "./Assets/adventure_content.xlsx"


#Files needed for this to work:
# - adventure_log.txt
#   = Results of the last x adventures where x = MAX_ADVENTURES
# - adventure_tracker.json
#   = The adventure queue, to make sure the adventures complete at the right time.
# - leaderboard.json
#   = For a planned future upgrade, when the program is made to sync with other instances on the network.
# - adventurer.json
#   = All the details about the adventurer -- stats, inventory, etc.
# - adventure_content.xlsx **this is the only one that the game won't automatically create
#   = Worldbuilding stuff. Has the following tabs:
#       ["Enemies",           #Enemy names/types sorted by difficulty
#        "Items",             #Consumables, their target stat, duration of effect, and magnitude
#        "Achievements",      #Idk what this is for yet.
#        "NarrativeEvents",   #Randomly dropped into the log
#        "SkillTree",         #For spending xp to upgrade the adventurer
#        "Weapons",           #Weapon names/types sorted by strength
#        "Armor"]             #Armor names/types sorted by strength


def get_content_tier(thing):
    #thing can be a weapon, armor, or enemy
    return (thing["type_index"] * 3 + thing["basename_index"] * 100) if thing else 0

def get_armor_strength(armor, content):
    max_armor_variability = 30 #was 3 before. Amounts to +/- a number up to this.
    if armor and content and content.get("Armor"):
        hp = (get_content_tier(armor)//100) * 5
        max_armor_variability = max(hp, max_armor_variability) #make sure armor never gives a negative value to newbies? idk, maybe reconsider this line.
        color = armor["ColorModifier"]
        color_index = content["Armor"]["ColorModifier"].index(color) % max_armor_variability
        variation = color_index+1
        #variation = round(5 * (1.0 + color_index * 0.05)) - 5  # ±0, ±1, ±2...
        return hp, variation
    else:
        return 0, 0

def get_weapon_strength(weapon, content):
    max_weapon_variability = 50 #was 5 before. Amounts to +/- a percentage of the weapon's hit.
    if weapon and content and content.get("Weapons"):
        weapon_base = (get_content_tier(weapon)//100) * 2
        weapon_color = weapon["ColorModifier"]
        color_index = content["Weapons"]["ColorModifier"].index(weapon_color) % 50
        variation = (color_index + 1)/ 100
        #print("-------------\nweapon tier:",get_content_tier(weapon),"\nweapon_color",weapon_color,"\ncolor_index",color_index,"\nvariation",variation,"\nweapon_base",weapon_base)
        return weapon_base, variation
    else:
        return 0, 0

class EnemyGenerator:
    def __init__(self, enemies_content, max_index=100):
        self.enemies = enemies_content
        self.max_index = max_index
        self.basename_weight = 100 #effective weight of the basename index
        self.stages_per_difficulty = STAGES_PER_DIFFICULTY
        self.difficulty_range = DIFFICULTY_RANGE 
        self.type_modifier_weight = 3  # From type_index * 3
        self.hp_difficulty_multiplier = 0.7  # From base_difficulty * 0.7
        self.attack_difficulty_multiplier = 0.07  # From base_difficulty * 0.07
        self.base_enemy_hp = 100  # Base HP constant
        self.base_enemy_attack = 7  # Base attack constant
        self.color_factors = [1.0 + (i % 5) * 0.05 for i in range(len(self.enemies["ColorModifier"]))]  # ±0%, ±5%, ±10%
        self.randomness_min = 0.8
        self.randomness_max = 1.2

    def generate(self, stage, is_boss = None):
        if is_boss is None:
            is_boss = (stage % STAGES_PER_BOSS == 0) and stage > 0
        boss_tier_multiplier = max(1,(stage//STAGES_PER_BOSS) * 1.8) #Bosses are are difficulty * 2. Enemies after a boss should back that off only a little bit.
        
        # Basenames as the major index, and types as the minor index. With 1/3 overlap between indices.
        # Effective index = 3 * difficulty / 2, where:
        # - integer part becomes basename index
        # - fractional part determines type index range
        
        # STEP 1: convert the stage to a difficulty level.
        # it is divided by max index so that difficulty level will affect types more than basenames
        current_difficulty = ((stage // self.stages_per_difficulty)) / self.max_index
        difficulty_high = (current_difficulty + self.difficulty_range) / self.max_index
        difficulty_low = (current_difficulty - self.difficulty_range) / self.max_index
        
        # STEP 2: Convert difficulty level to high/low type and basename indices
        # get actual high and low
        range_high = 3*difficulty_high/2
        range_low = (3*difficulty_low/2)-1
        range_mid = (3*current_difficulty/2) - 0.5
        #basenames are just the range without the fraction
        basename_high = int(range_high)
        basename_low = int(range_low)
        basename_mid = int(range_mid)
        #subtract to get just the fraction. That decides how far into the types we go.
        type_high = (range_high - basename_high) * self.max_index
        type_low = (range_low - basename_low) * self.max_index
        type_mid = (range_mid - basename_low) * self.max_index

        # STEP 3: select a random index within the range:
        basename_index = random.triangular(basename_high, basename_low, basename_mid)
        type_index = random.triangular(type_high, type_low, type_mid)

        # Clamp indices just in case
        basename_index = round(min(max(0, basename_index), self.max_index - 1))
        type_index = round(min(max(0, type_index), self.max_index - 1))

        # STEP 4: Choose a random color
        color_index = random.randint(0, len(self.color_factors) - 1)
        color_mod = self.enemies["ColorModifier"][color_index]
        color_factor = self.color_factors[color_index]
        effective_color_factor = random.uniform(0.95, color_factor)

        #get the names
        type_mod = self.enemies["TypeModifier"][type_index]
        bn_column = random.choice(["BaseName", "BaseName2", "BaseName3"])
        base_name = self.enemies[bn_column][basename_index]
        enemy_name = f"Boss {type_mod} {color_mod} {base_name}" if is_boss else f"{type_mod} {color_mod} {base_name}"

        # STEP 5: Stats
        base_difficulty = ((type_index + 1) * self.type_modifier_weight + basename_index * self.basename_weight) * effective_color_factor
        base_difficulty = (base_difficulty + 1) * (1 + stage * DIFFICULTY_INCREASE_PER_STAGE)

        base_hp = base_difficulty * self.hp_difficulty_multiplier + self.base_enemy_hp * boss_tier_multiplier + stage * 3
        base_attack = (base_difficulty * self.attack_difficulty_multiplier + self.base_enemy_attack) * (1 + boss_tier_multiplier) + stage * 2

        if is_boss:
            base_hp *= 2
            base_attack *= 2
        
        return {
            "name": enemy_name,
            "hp": int(base_hp * random.uniform(self.randomness_min, self.randomness_max)),
            "attack": int(base_attack * random.uniform(self.randomness_min, self.randomness_max)),
            "type_index": type_index,
            "basename_index": basename_index
        }

    def get_max_hp(self, stage):
        # Max HP at stage (max color_factor, max randomness)
        is_boss = (stage % STAGES_PER_BOSS == 0) and stage > 0
        boss_tier_multiplier = max(1,(stage//STAGES_PER_BOSS) * 1.8)
        current_difficulty = ((stage // self.stages_per_difficulty)) / self.max_index
        difficulty_high = (current_difficulty + self.difficulty_range) / self.max_index
        range_high = 3*difficulty_high/2
        basename_index = int(range_high)
        type_index = (range_high - basename_index) * self.max_index
        basename_index = round(min(max(0, basename_index), self.max_index - 1))
        type_index = round(min(max(0, type_index), self.max_index - 1))
        #i actually don't remember the max color factor. That calc is weird.
        color_index = random.randint(0, len(self.color_factors) - 1)
        color_mod = self.enemies["ColorModifier"][color_index]
        color_factor = self.color_factors[color_index]
        effective_color_factor = random.uniform(0.95, color_factor)
        base_difficulty = ((type_index + 1) * self.type_modifier_weight + basename_index * self.basename_weight) * effective_color_factor
        base_difficulty = (base_difficulty + 1) * (1 + stage * DIFFICULTY_INCREASE_PER_STAGE)
        base_hp = base_difficulty * self.hp_difficulty_multiplier + self.base_enemy_hp * boss_tier_multiplier + stage * 3
        if is_boss:
            base_hp *= 2
        return base_hp * self.randomness_max

class Adventurer:
    def __init__(self, name="Boxcar", level=1, xp=0, base_hp=80, base_attack=10, inventory=None, skills=None, 
                 achievements=None, equipped_weapon=None, equipped_armor=None, id=str(uuid.uuid4()), tasks_completed=0,
                 enemy_defeats=0, achievement_progress=None, achievements_awarded=None, narrative_progress=0, last_narrative_date=None,
                 recent_items=None, pending_used_items=None, saved_items=None, mission_inventory=None):
        self.id = id
        self.name = name
        self.level = level
        self.xp = xp
        self.base_hp = base_hp
        self.base_attack = base_attack
        self.inventory = inventory or []
        self.skills = skills or []
        self.achievements = achievements or []
        self.equipped_weapon = equipped_weapon
        self.equipped_armor = equipped_armor
        self.tasks_completed = tasks_completed
        self.enemy_defeats = enemy_defeats
        self.achievement_progress = achievement_progress or {}
        self.achievements_awarded = achievements_awarded or {}
        self.narrative_progress = narrative_progress
        self.recent_items = recent_items or [] #"Stored Items" go to self.inventory
        self.pending_used_items = pending_used_items or []
        self.saved_items = saved_items or []
        #self.mission_inventory = mission_inventory or []
        if last_narrative_date is None:
            self.last_narrative_date = datetime.datetime.now().isoformat()
        else:
            self.last_narrative_date = last_narrative_date

    def get_max_hp(self, armor=None, content=None): 
        if armor is None:
            armor = self.equipped_armor
        armor_HP, variation = get_armor_strength(armor, content)
        max_hp = self.base_hp + armor_HP# + variation #variation removed because it averages out to base in theory, and this function is used for updating the hp icon in the main program.
        return max_hp

    def get_adventure_start_hp(self, armor=None, content=None):
        if armor is None:
            armor = self.equipped_armor
        armor_HP, variation = get_armor_strength(armor, content)
        effective_hp = self.base_hp + armor_HP + round(random.uniform(0-variation, 0+variation))
        return effective_hp

    def get_attack(self, weapon=None, content=None):
        if weapon is None:
            weapon = self.equipped_weapon
        weapon_damage, variation = get_weapon_strength(weapon, content)
        damage = self.base_attack + weapon_damage
        multiplier = random.uniform(1 - variation, 1 + variation)
        #print("----------\nmultiplier",multiplier,"\ndamage",damage)
        return damage, multiplier

    def is_better_gear(self, new_gear, current_gear, gear_type, content):
        if not new_gear or not content[gear_type]:
            return False
        if not current_gear:
            return True
        # Gear uses type and base with strength = type_index * 3 + base_index * 100.
        # This is intended to make the gear types have some overlapping strength ratings
#        new_index = (new_gear["type_index"], new_gear["basename_index"])
#        current_index = (current_gear["type_index"], current_gear["basename_index"])
        new_difficulty = get_content_tier(new_gear)
        current_difficulty = get_content_tier(current_gear)
        if new_difficulty > current_difficulty:
            return True
        if new_difficulty == current_difficulty and new_gear["ColorModifier"] != current_gear["ColorModifier"]:
            return True
        return False

    def level_up(self):
        xp_needed = 3000 + (self.level * 10)  # XP per level
        while self.xp >= xp_needed:
            self.level += 1
            self.xp -= xp_needed   # Subtract XP for level-up
            self.base_hp += 20
            self.base_attack += 5
            xp_needed = 3000 + (self.level * 10)
        return self.level

class AdventureManager:
    global MAX_ADVENTURES, STAGES_PER_DIFFICULTY, DIFFICULTY_RANGE
    def __init__(self, task_manager):
        if not Adventure_Feature_Enabled:
            self.task_manager = None
            self.adventurer = None
            self.leaderboard = None
            self.content = None
            self.log_file = None
            return
        self.task_manager = task_manager
        self.adventurer = self.load_adventurer()
        self.leaderboard = self.load_leaderboard()
        self.content = self.load_content()
        #    def __init__(self, enemies_content, max_index=100, difficulty_constant=100, stages_per_difficulty=5, difficulty_range=10):
        self.enemy_generator = EnemyGenerator(
            self.content["Enemies"],
            max_index=len(self.content["Enemies"]["TypeModifier"])
        )
        self.log_file = LOG_FILE_PATH
        self.tracker_file = TRACKER_FILE_PATH
        self.current_timer_id = None
        self.completed_task_ids = []
        self.adventure_queue = []
        self.stats_window = None
        self.current_adventure_data = None
        self.hp_label = None
        self.load_tracker()
        

    def set_hp_label(self, hp_label):
        #called once when the main window is initializing in setup_gui
        #moved to its own function because this label has to be created before this stuff can happen.
        if not Adventure_Feature_Enabled:
            return
        self.hp_label = hp_label
        if self.hp_label:
            self.update_current_hp()
            self.initialize_adventures()

    def load_adventurer(self):
        try:
            with open(ADVENTURER_FILE_PATH, "r") as f:
                data = json.load(f)
                return Adventurer(**data)
        except FileNotFoundError:
            adventurer = Adventurer()
            with open(ADVENTURER_FILE_PATH, "w") as f:
                json.dump(vars(adventurer), f, default=str)
            return adventurer

    def save_adventurer(self):
        with open(ADVENTURER_FILE_PATH, "w") as f:
            json.dump(vars(self.adventurer), f, default=str)

    def load_leaderboard(self):
        try:
            with open(LEADERBOARD_FILE_PATH, "r") as f:
                data = json.load(f)
                return data.get("leaderboard", [{"name": self.adventurer.name, "xp": self.adventurer.xp}])
        except FileNotFoundError:
            leaderboard = [{"name": self.adventurer.name, "xp": self.adventurer.xp}]
            with open(LEADERBOARD_FILE_PATH, "w") as f:
                json.dump({"leaderboard": leaderboard}, f, default=str)
            return leaderboard

    def save_leaderboard(self):
        try:
            with open(LEADERBOARD_FILE_PATH, "r") as f:
                data = json.load(f)
        except FileNotFoundError:
            data = {}
        data["leaderboard"] = self.leaderboard
        with open(LEADERBOARD_FILE_PATH, "w") as f:
            json.dump(data, f, default=str)

    def load_content(self):
        wb = openpyxl.load_workbook(CONTENT_FILE_PATH)
        content = {}
        for sheet_name in ["Enemies", "Items", "Achievements", "NarrativeEvents", "SequentialNarrative", "SkillTree", "Weapons", "Armor"]:
            sheet = wb[sheet_name]
            if sheet_name in ["Weapons", "Armor"]:
                content[sheet_name] = {
                    "TypeModifier": [],
                    "ColorModifier": [], 
                    "BaseName": []
                }
            elif sheet_name in ["Enemies"]:
                content[sheet_name] = {
                    "TypeModifier": [],
                    "ColorModifier": [],
                    "BaseName": [],
                    "BaseName2": [], #The user will have between 6 and 9 possible basenames for enemies, depending on level
                    "BaseName3": []
                }
            else:
                content[sheet_name] = []
            
            headers = [cell.value for cell in sheet[1]]
            if sheet_name == "Achievements":
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    if row_dict.get("Effect Stat") and "," in row_dict["Effect Stat"]:
                        row_dict["Effect Stat"] = row_dict["Effect Stat"].split(",")
                        row_dict["Effect"] = [int(x) if x.isdigit() else x for x in row_dict["Effect"].split(",")]
                    else:
                        row_dict["Effect Stat"] = [row_dict.get("Effect Stat") or "None"]
                        row_dict["Effect"] = [row_dict.get("Effect") or "None"]
                    content[sheet_name].append(row_dict)
            elif sheet_name in ["Weapons", "Armor"]:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    for key in ["TypeModifier", "ColorModifier", "BaseName"]:
                        if row_dict[key] and row_dict[key] not in content[sheet_name][key]:
                            content[sheet_name][key].append(row_dict[key])
            elif sheet_name in ["Enemies"]:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    for key in ["TypeModifier", "ColorModifier", "BaseName", "BaseName2", "BaseName3"]:
                        if row_dict[key] and row_dict[key] not in content[sheet_name][key]:
                            content[sheet_name][key].append(row_dict[key])
            else:
                for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
                    row_dict = dict(zip(headers, row))
                    content[sheet_name].append(row_dict)
        return content

    def load_tracker(self):
        try:
            with open(self.tracker_file, "r") as f:
                data = json.load(f)
                self.completed_task_ids = [
                    {
                        "task_id": entry["task_id"],
                        "date": datetime.datetime.fromisoformat(entry["date"])
                    }
                    for entry in data.get("completed_task_ids", [])
                ]
                self.adventure_queue = [
                    {
                        "task_id": entry["task_id"],
                        "priority": entry["priority"],
                        "completion_date": datetime.datetime.fromisoformat(entry["completion_date"]),
                        "end_time": datetime.datetime.fromisoformat(entry["end_time"]),
                        "start_time": datetime.datetime.fromisoformat(entry.get("start_time", entry["end_time"])),
                        "log": entry.get("log", []),
                        "hp_changes": entry.get("hp_changes", []),
                        "battle_times": [datetime.datetime.fromisoformat(bt) for bt in entry.get("battle_times", [])],
                        "temp_state": entry.get("temp_state", {
                            "xp": 0,
                            "inventory": [],
                            "equipped_weapon": self.adventurer.equipped_weapon,
                            "equipped_armor": self.adventurer.equipped_armor,
                            "enemy_defeats": 0,
                            "achievement_progress": {},
                            "tasks_completed": 0
                        }),
                        "short_desc": entry.get("short_desc", ""),
                        "is_win": entry.get("is_win",False)
                    }
                    for entry in data.get("adventure_queue", [])
                ]
        except:
            self.completed_task_ids = []
            self.adventure_queue = []
            with open(self.tracker_file, "w") as f:
                json.dump({
                    "completed_task_ids": [],
                    "adventure_queue": []
                }, f, default=str)

    def save_tracker(self):
        data = {
            "completed_task_ids": [
                {
                    "task_id": entry["task_id"],
                    "date": entry["date"].isoformat()
                }
                for entry in self.completed_task_ids
            ],
            "adventure_queue": [
                {
                    "task_id": entry["task_id"],
                    "priority": entry["priority"],
                    "completion_date": entry["completion_date"].isoformat(),
                    "end_time": entry["end_time"].isoformat(),
                    "start_time": entry["start_time"].isoformat(),
                    "log": entry["log"],
                    "hp_changes": entry["hp_changes"],
                    "battle_times": [bt.isoformat() for bt in entry["battle_times"]],
                    "temp_state": {
                        "xp": entry["temp_state"]["xp"],
                        "inventory": entry["temp_state"]["inventory"],
                        "equipped_weapon": entry["temp_state"]["equipped_weapon"],
                        "equipped_armor": entry["temp_state"]["equipped_armor"],
                        "enemy_defeats": entry["temp_state"]["enemy_defeats"],
                        "achievement_progress": entry["temp_state"]["achievement_progress"],
                        "tasks_completed": entry["temp_state"]["tasks_completed"]
                    },
                    "short_desc": entry["short_desc"],
                    "is_win": entry["is_win"]
                }
                for entry in self.adventure_queue
            ]
        }
        with open(self.tracker_file, "w") as f:
            json.dump(data, f, default=str)

    def initialize_adventures(self):
        now = datetime.datetime.now()
        today_midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
        
        self.completed_task_ids = [
            entry for entry in self.completed_task_ids
            if entry["date"] >= today_midnight
        ]
                                             
        overdue = [
            entry for entry in self.adventure_queue
            if entry["end_time"] <= now
        ]
        self.adventure_queue = [
            entry for entry in self.adventure_queue
            if entry["end_time"] > now
        ]
        
        for adventure in overdue:
            # Run adventure to compute results (no timer needed)
            log, hp_changes, temp_state = self.run_adventure(
                adventure["priority"],
                adventure["completion_date"],
                adventure["task_id"],
                adventure["short_desc"],
                is_win = adventure["is_win"]
            )
            # Update adventure with results
            adventure["log"] = log
            adventure["hp_changes"] = hp_changes
            adventure["temp_state"] = temp_state
            adventure["battle_times"] = [now + datetime.timedelta(seconds=i * (adventure["end_time"] - adventure["start_time"]).total_seconds() / len(hp_changes)) for i in range(len(hp_changes))] if hp_changes else []
            self.complete_adventure(adventure)

        # Prorate ongoing adventures
        for adventure in self.adventure_queue:
            if "start_time" not in adventure:
                adventure["start_time"] = now
            # Use stored results if available, else compute
            if not adventure.get("hp_changes"):
                log, hp_changes, temp_state = self.run_adventure(
                    adventure["priority"],
                    adventure["completion_date"],
                    adventure["task_id"],
                    adventure["short_desc"],
                    is_win = adventure["is_win"]
                )
                adventure["log"] = log
                adventure["hp_changes"] = hp_changes
                adventure["temp_state"] = temp_state
                adventure["battle_times"] = [now + datetime.timedelta(seconds=i * (adventure["end_time"] - adventure["start_time"]).total_seconds() / len(hp_changes)) for i in range(len(hp_changes))] if hp_changes else []
        
        self.start_next_adventure()

    def queue_adventure(self, task_priority, completion_date, task_id, short_desc, is_win):
        if not Adventure_Feature_Enabled:
            return
        now = datetime.datetime.now()
        today_midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
        
        if any(
            entry["task_id"] == task_id and entry["date"] >= today_midnight
            for entry in self.completed_task_ids
        ):
            return
        
        self.completed_task_ids.append({
            "task_id": task_id,
            "date": now
        })
        
        duration_seconds = 15 * self.adventurer.level
        if duration_seconds > 14400: #seconds in 4 hrs (starts to matter at about lvl 1000)
            duration_seconds = 14400 #mission duration limited so that 6 missions/day can always be completed
        duration = datetime.timedelta(seconds=duration_seconds)
        
        if not self.adventure_queue:
            end_time = now + duration
        else:
            last_end_time = max(
                entry["end_time"] for entry in self.adventure_queue
            )
            end_time = last_end_time + duration

        log, hp_changes, temp_state = self.run_adventure(task_priority, completion_date, task_id, short_desc, is_win = is_win)
        total_time = (end_time - now).total_seconds()
        battle_count = len(hp_changes)
        adventure = {
            "task_id": task_id,
            "priority": task_priority,
            "is_win": is_win,
            "completion_date": completion_date,
            "end_time": end_time,
            "start_time": now,
            "log": log,
            "hp_changes": hp_changes,
            "battle_times": [now + datetime.timedelta(seconds=i * total_time / battle_count) for i in range(battle_count)] if battle_count > 0 else [],
            "temp_state": temp_state,
            "short_desc": short_desc
        }
        self.adventure_queue.append(adventure)
        
        self.save_tracker()
        self.start_next_adventure()
        
    def start_next_adventure(self):
        if not self.adventure_queue: 
            #no adventures in queue. Indicate no adventure.
            self.current_adventure_data = None
            #update_current_hp will set the hp indicator to max hp without asterisk.
            self.update_current_hp()
            return
        
        #Capture the current time.
        now = datetime.datetime.now()
        #"next adventure" refers to the adventure which this function has been called to start
        # it is the one with the nearest "end time".
        next_adventure = min(
            self.adventure_queue,
            key=lambda x: x["end_time"]
        )
        #print(next_adventure)
        
        time_remaining = (next_adventure["end_time"] - now).total_seconds() * 1000
        if time_remaining <= 0:
            #if the adventure ended in the past, then just complete it.
            self.complete_adventure(next_adventure)
        else:
            #if the adventure ends in the future, set it as the current adventure, and start the update_hp timers.
            if not self.current_adventure_data or self.current_adventure_data["task_id"] != next_adventure["task_id"]:
                self.current_adventure_data = {
                    "hp_changes": next_adventure["hp_changes"],
                    "battle_times": next_adventure["battle_times"],
                    "start_time": next_adventure["start_time"],
                    "end_time": next_adventure["end_time"],
                    "task_id": next_adventure["task_id"]
                }
            
            #at the end of the update_hp timer sequence, it will complete_adventure and then call this function again.
            self.update_current_hp()

    def complete_adventure(self, adventure):
        # The call structure here is...
        # queue():
        #   adventure = run()
        #   start_next():
        #       complete(adventure)
        #       update_current_hp(adventure)
        log = adventure["log"]
        hp_changes = adventure["hp_changes"]
        temp_state = adventure["temp_state"]
        
        #remove this adventure from the queue
        self.adventure_queue = [
            entry for entry in self.adventure_queue
            if entry["task_id"] != adventure["task_id"]
        ]
        
        # Apply stat changes from temp_state to the adventurer.
        self.adventurer.xp += temp_state["xp"]
        self.adventurer.inventory.extend(self.adventurer.recent_items)
        self.adventurer.recent_items = temp_state["inventory"]
        self.adventurer.equipped_weapon = temp_state["equipped_weapon"]
        self.adventurer.equipped_armor = temp_state["equipped_armor"]
        self.adventurer.enemy_defeats += temp_state["enemy_defeats"]
        self.adventurer.achievement_progress.update(temp_state["achievement_progress"])
        self.adventurer.tasks_completed += temp_state["tasks_completed"]
        
        #do any level-ups that need to be done
        current_level = self.adventurer.level
        self.adventurer.level_up()
        if self.adventurer.level > current_level:
            log.append(f"{self.adventurer.name} leveled up to Level {self.adventurer.level}!")
        
        #update the leaderboard (partial feature)
        for entry in self.leaderboard:
            if entry["name"] == self.adventurer.name:
                entry["xp"] = self.adventurer.xp
                break
        else:
            self.leaderboard.append({"name": self.adventurer.name, "xp": self.adventurer.xp})
        self.leaderboard.sort(key=lambda x: x["xp"], reverse=True)
        
        #save results to files
        self.save_adventurer()
        self.save_leaderboard()
        
        # Write to log
        self.prune_and_prepend_log(log)
        #self.prune_log()
        #with open(self.log_file, "a", encoding="utf-8") as f:
        #    f.write("\n".join(log) + "\n\n")
        
        #clear the current adventure.
        self.current_adventure_data = None
        #remove current adventure from tracker
        self.save_tracker()
        
    def get_sequential_narrative_event(self):
        # ascii flourish:
        flourish_top =      "╭═════════════════════════ ༺ ✦ ༻ ═════════════════════════╮\n║♦ STORY EVENT"
        flourish_body =     "\n║♦ "
        flourish_bottom =   "\n╰═════════════════════════ ༺ ✦ ༻ ═════════════════════════╯"
        
        # Check if narrative events exist
        if not self.content.get("SequentialNarrative"):
            print("FAILED TO GET SEQUENTIAL NARRATIVE EVENT")
            return None
        
        # Handle out-of-bounds progress
        max_events = len(self.content["SequentialNarrative"])
        if self.adventurer.narrative_progress >= max_events:
            self.adventurer.narrative_progress = 0  # Reset to start (or stop, if preferred)
            self.save_adventurer()
        
        try:
            # Get the current event
            event = self.content["SequentialNarrative"][self.adventurer.narrative_progress]["Description"]
            # Increment progress for next event
            self.adventurer.narrative_progress += 1
            #self.save_adventurer() #might not be necessary, (might even be disruptive), because we save at the end of the adventure
            # Format with flourish (center event text in 51-char width)
            wrapped_event = textwrap.fill(event, width=50)
            centered_event = [line.center(50) for line in wrapped_event.split('\n')]
            formatted_event = f"{flourish_top}{flourish_body}{flourish_body.join(centered_event)}{flourish_bottom}"
            
            return formatted_event
        except Exception as e:
            # Fallback for missing events or malformed data
            print("FAILED TO GET SEQUENTIAL NARRATIVE EVENT:", e)
            return None
    
    def update_current_hp(self):
        # Updates the label on the main task window
        # Also sets timers to continuously update it "during" adventures.
        if not self.hp_label:
            #the first time this is called during initialization, the hp_label isn't instantiated by the main program yet.
            return
            
        if not self.current_adventure_data:
            #this is the case when queue_adventure calls with no active adventures.
            max_hp = self.adventurer.get_max_hp(None,self.content)
            self.hp_label.config(text=f"H: {max_hp}")
            return
        
        now = datetime.datetime.now()
        hp_changes = self.current_adventure_data["hp_changes"]
        battle_times = self.current_adventure_data["battle_times"]
        start_time = self.current_adventure_data["start_time"]
        this_id = self.current_adventure_data["task_id"]
        this_adventure = next((entry for entry in self.adventure_queue if entry.get("task_id") == this_id), None)
        
        max_hp = self.adventurer.get_max_hp(None,self.content)
        current_hp = max_hp
        
        #real-talk I'm not sure how this part works.
        for i, battle_time in enumerate(battle_times):
            if now >= battle_time and i < len(hp_changes):
                current_hp = hp_changes[i]
        
        #update the hp label on the main display
        current_hp = min(current_hp, max_hp)
        self.hp_label.config(text=f"*H: {current_hp}")
        
        #if the battle isn't over yet, set the timer and call again.
        if now < self.current_adventure_data["end_time"]:
            self.task_manager.root.after(100, self.update_current_hp)
        else:
            #if we reached the end of the battle, start the next adventure.
            if this_adventure:
                self.complete_adventure(this_adventure)
            max_hp = self.adventurer.get_max_hp(None,self.content)
            self.hp_label.config(text=f"H: {max_hp}")
            self.start_next_adventure()

    def use_recent_item(self, item_type):
        #recent items are dictionaries of the item's attributes..... maybe will fix that later.
        items = [item for item in self.adventurer.recent_items if next((i for i in self.content["Items"] if i["ItemName"] == item), None)["TargetStat"] == item_type]
        if not items:
            messagebox.showinfo("No Items", f"No {item_type} items in recent inventory.")
            return
        item_name = random.choice(items)
        item = next(i for i in self.content["Items"] if i["ItemName"] == item_name)
        self.adventurer.recent_items.remove(item_name)
        self.adventurer.pending_used_items.append(item)
        if item["TargetStat"] == "HP":
            self.adventurer.base_hp += item["Effect"]
            self.adventurer.base_hp = max(0, min(self.adventurer.base_hp, 1000))
        elif item["TargetStat"] == "Attack":
            self.adventurer.base_attack += item["Effect"]
            self.adventurer.base_attack = max(0, min(self.adventurer.base_attack, 1000))
        self.save_adventurer()
        self.refresh_adventurer_window(self.stats_window.winfo_children()[0] if self.stats_window else None)

    def use_stored_item(self, item_type):
        #stored items are the string name of the item.
        items = [item for item in self.adventurer.inventory if next((i for i in self.content["Items"] if i["ItemName"] == item), None)["TargetStat"] == item_type]
        if not items:
            messagebox.showinfo("No Items", f"No {item_type} items in stored inventory.")
            return
        item_name = random.choice(items)
        item = next(i for i in self.content["Items"] if i["ItemName"] == item_name)
        self.adventurer.inventory.remove(item_name)
        self.adventurer.pending_used_items.append(item)
        if item["TargetStat"] == "HP":
            self.adventurer.base_hp += item["Effect"]
            self.adventurer.base_hp = max(0, min(self.adventurer.base_hp, 1000))
        elif item["TargetStat"] == "Attack":
            self.adventurer.base_attack += item["Effect"]
            self.adventurer.base_attack = max(0, min(self.adventurer.base_attack, 1000))
        self.save_adventurer()
        self.refresh_adventurer_window(self.stats_window.winfo_children()[0] if self.stats_window else None)

    def use_saved_item(self, item_type):
        #stored items are the string name of the item.
        items = [item for item in self.adventurer.saved_items if next((i for i in self.content["Items"] if i["ItemName"] == item), None)["TargetStat"] == item_type]
        if not items:
            messagebox.showinfo("No Items", f"No {item_type} items in saved inventory.")
            return
        item_name = random.choice(items)
        item = next(i for i in self.content["Items"] if i["ItemName"] == item_name)
        self.adventurer.saved_items.remove(item_name)
        self.adventurer.pending_used_items.append(item)
        if item["TargetStat"] == "HP":
            self.adventurer.base_hp += item["Effect"]
            self.adventurer.base_hp = max(0, min(self.adventurer.base_hp, 1000))
        elif item["TargetStat"] == "Attack":
            self.adventurer.base_attack += item["Effect"]
            self.adventurer.base_attack = max(0, min(self.adventurer.base_attack, 1000))
        self.save_adventurer()
        self.refresh_adventurer_window(self.stats_window.winfo_children()[0] if self.stats_window else None)

    def trash_recent_item(self, item_type):
        items = [item for item in self.adventurer.recent_items if next((i for i in self.content["Items"] if i["ItemName"] == item), None)["TargetStat"] == item_type]
        if not items:
            messagebox.showinfo("No Items", f"No {item_type} items to trash.")
            return
        item_name = random.choice(items)
        self.adventurer.recent_items.remove(item_name)
        self.save_adventurer()
        self.refresh_adventurer_window(self.stats_window.winfo_children()[0] if self.stats_window else None)
        
    
    def trash_stored_item(self, item_type):
        items = [item for item in self.adventurer.inventory if next((i for i in self.content["Items"] if i["ItemName"] == item), None)["TargetStat"] == item_type]
        if not items:
            messagebox.showinfo("No Items", f"No {item_type} items to trash.")
            return
        item_name = random.choice(items)
        self.adventurer.inventory.remove(item_name)
        self.save_adventurer()
        self.refresh_adventurer_window(self.stats_window.winfo_children()[0] if self.stats_window else None)
    
    def trash_saved_item(self, item_type):
        items = [item for item in self.adventurer.saved_items if next((i for i in self.content["Items"] if i["ItemName"] == item), None)["TargetStat"] == item_type]
        if not items:
            messagebox.showinfo("No Items", f"No {item_type} items to trash.")
            return
        item_name = random.choice(items)
        self.adventurer.saved_items.remove(item_name)
        self.save_adventurer()
        self.refresh_adventurer_window(self.stats_window.winfo_children()[0] if self.stats_window else None)
    
    def save_stored_item(self, item_type):
        items = [item for item in self.adventurer.inventory if next((i for i in self.content["Items"] if i["ItemName"] == item), None)["TargetStat"] == item_type]
        if not items:
            messagebox.showinfo("No Items", f"No {item_type} items to trash.")
            return
        item_name = random.choice(items)
        self.adventurer.inventory.remove(item_name)
        self.adventurer.saved_items.append(item_name)
        self.save_adventurer()
        self.refresh_adventurer_window(self.stats_window.winfo_children()[0] if self.stats_window else None)

    def run_adventure(self, task_priority, completion_date, task_id, short_desc, is_win = False):
        log = []
        hp_changes = []
        #store the results of the battle in temp_state so that we can apply the changes after all the timers.
        temp_state = {
            "xp": 0,
            "inventory": [],
            "equipped_weapon": self.adventurer.equipped_weapon,
            "equipped_armor": self.adventurer.equipped_armor,
            "enemy_defeats": 0,
            "achievement_progress": {},
            "tasks_completed": 0
        }
        max_hp = self.adventurer.get_adventure_start_hp(None, self.content)
        current_hp = max_hp
        consumable_effects = {"HP": 0, "Attack": 0}
        
        #backdating a task completion loses xp.
        today = datetime.datetime.now().date()
        days_backdated = (today - completion_date.date()).days
        xp_multiplier = max(0.2, 1 - 0.5 * days_backdated)
        effective_priority = (task_priority - 100) / 2 if task_priority > 100 else task_priority
        xp_from_task = effective_priority * 10 * xp_multiplier
        temp_state["xp"] = xp_from_task
        
        killcount = 0 #enemies killed this adventure
        stage = 0 #combined enemies and events. Used for enemy difficulty
        xp_from_enemies = 0
        
        adventure_start_note = random.choice(["is on a quest to", "seeks to", "ventures forth to", "aims to", "resolves to", "must"])
        
        log.append(f"╭───────────────────── ✧ ✦ ✧ ─────────────────────╮")
        log.append(f"[{datetime.datetime.now()}] Adventure begins for {self.adventurer.name} (Level {self.adventurer.level}, HP: {current_hp}, Attack: {self.adventurer.get_attack(None, self.content)[0]})")
        log.append(f"{self.adventurer.name} {adventure_start_note} {short_desc}")
        log.append(f"╰───────────────────────────────────────────────────╯")

        #Calculate initial consumable buffs
        if self.adventurer.pending_used_items:
            log.append("Used items:")
            for item in self.adventurer.pending_used_items:
                log.append(f"  - {item['ItemName']}: {item['TargetStat']} {'+' if item['Effect'] >= 0 else ''}{item['Effect']}")
                consumable_effects[item["TargetStat"]] += item["Effect"]
            self.adventurer.pending_used_items = []
            self.save_adventurer()
        
        current_hp += consumable_effects["HP"]
        hp_changes.append(current_hp) #let the user sit a little at full health before changing his HP right away
        



        # If the hero survives his consumables, calculate stages to skip (one-shot enemies)
        if current_hp > 0:
            player_attack, multiplier = self.adventurer.get_attack(None, self.content)
            attack = (player_attack * multiplier) + consumable_effects["Attack"]
            if (attack < 0):
                attack = 1
            
            skipped_stages = 0
            if attack > self.enemy_generator.generate(0)["hp"]:
                print("player is skipping stages with attack:",attack)
                # Binary search for max stage where attack >= max_hp
                low, high = 0, HIGHEST_STAGE
                while low <= high:
                    mid = (low + high) // 2
                    if attack >= self.enemy_generator.generate(mid)["hp"]:
                        skipped_stages = mid
                        low = mid + 1
                    else:
                        high = mid - 1
                    print("low, mid, high", low, mid, high)
            if skipped_stages >= HIGHEST_STAGE-1:
                log.append(f"═───────◎───────══✦══───────◎───────═")
                log.append(f"{self.adventurer.name} has conquered the game.")
                log.append(f"═───────◎───────══✧══───────◎───────═")
            if skipped_stages > 0:
                temp_state["enemy_defeats"] += skipped_stages
                #temp_state["achievement_progress"]["kills"] = temp_state["achievement_progress"].get("kills", 0) + skipped_stages
                destruction_word = random.choice([
                    "obliterated",
                    "annihilated",
                    "eradicated",
                    "eliminated",
                    "vaporized",
                    "exterminated",
                    "neutralized",
                    "disintegrated",
                    "terminated"
                ])
                log.append(f"═════🔥 {self.adventurer.name} {destruction_word} {skipped_stages} weak enemies instantly! 🔥═════")
            
            stage = skipped_stages
            killcount = stage

        if days_backdated > 0:
            temp_state["achievement_progress"]["backdated_tasks"] = temp_state["achievement_progress"].get("backdated_tasks", 0) + 1
        
        encounter_count = 0
        while current_hp > 0 and encounter_count < MAX_ENCOUNTERS:
            # Apply consumables
            consumable_effects["HP"] = 0

            
            #decide what kind of encounter it is
            # "end_time": entry["end_time"].isoformat(), "date": datetime.datetime.fromisoformat(entry["date"])
            if random.random() < 0.07: #EVENT CHANCE
                sequential_event_success = False
                last_narrative_date = datetime.datetime.fromisoformat(self.adventurer.last_narrative_date)
                days_since_last_story = (datetime.datetime.now().date() - last_narrative_date.date()).days
                story_chance = (days_since_last_story/10)
                #print("Chance of sequential story:",story_chance)
                if (days_since_last_story > 1) and (random.random() < story_chance): #extra-rare narrative event. max 1/day
                    sequential_event = self.get_sequential_narrative_event()
                    if not (sequential_event is None):
                        log.append(sequential_event)
                        sequential_event_success = True
                        temp_state["achievement_progress"]["sequential_narrative"] = temp_state["achievement_progress"].get("sequential_narrative", 0) + 1
                        self.adventurer.last_narrative_date = datetime.datetime.now().isoformat()
                if not sequential_event_success: #regular narrative/ambiance event
                    log.append(f"  ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
                    event = random.choice(self.content["NarrativeEvents"])
                    log.append(f" ✦ Event: {event['Description']}")
                hp_changes.append(current_hp)
                stage += 1
                continue
            
            log.append(f"  ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
            enemy = self.enemy_generator.generate(stage)
            is_boss = enemy['name'].startswith("Boss")
            log.append(f"> {'Confronted the Terrible' if is_boss else 'Encountered'} {enemy['name']} (HP: {enemy['hp']}, Attack: {enemy['attack']})")

            #battles! Adventurer attacks first
            while enemy["hp"] > 0 and current_hp > 0:
                player_damage, multiplier = self.adventurer.get_attack(None, self.content)
                hit_randomizer = random.uniform(1 - NORMAL_HIT_VARIATION, 1 + NORMAL_HIT_VARIATION)
                damage = int(((player_damage * multiplier) + consumable_effects["Attack"])*hit_randomizer)
                if damage < 0:
                    damage = 1
                #print("Player Damage:",player_damage,"\nmultiplier",multiplier,"\nconsumable_effects",consumable_effects["Attack"])
                #"Double Strike" is a sample skill...
                if "Double Strike" in self.adventurer.skills and random.random() < 0.2:
                    damage *= 2
                    log.append(f"   Double Strike!")
                    
                
                enemy["hp"] = int(enemy["hp"] - damage)
                log.append(f" - {self.adventurer.name} deals {damage} damage. Enemy HP: {enemy['hp']}")
                

                #dead enemy: calculate rewards
                if enemy["hp"] <= 0:
                    log.append(f"☑ {enemy['name']} defeated! +{XP_PER_KILL}xp")
                    xp_from_enemies = xp_from_enemies + XP_PER_KILL
                    killcount += 1
                    temp_state["xp"] = temp_state["xp"] + XP_PER_KILL
                    temp_state["enemy_defeats"] = temp_state["enemy_defeats"] + 1
                    temp_state["achievement_progress"]["kills"] = temp_state["achievement_progress"].get("kills", 0) + 1
                    
                    encounter_count += 1
                    stage += 1
                    # Gear drop (30% chance)
                    #as player levels up, the gear drop chance decreases and the consumable drop chance increases
                    if is_boss:
                        gear_drop_prob = 0.9
                        consumable_drop_prob = 0.5
                    elif self.adventurer.level <= 10:
                        gear_drop_prob =  0.1
                        consumable_drop_prob = 0.1
                    elif self.adventurer.level <= 20:
                        gear_drop_prob =  0.08
                        consumable_drop_prob = 0.2
                    elif self.adventurer.level <= 40:
                        gear_drop_prob =  0.05
                        consumable_drop_prob = 0.3
                    elif self.adventurer.level <= 100:
                        gear_drop_prob =  0.05
                        consumable_drop_prob = 0.3
                    elif self.adventurer.level <= 200:
                        gear_drop_prob =  0.05
                        consumable_drop_prob = 0.4
                    else:
                        gear_drop_prob = 0.05
                        consumable_drop_prob = 0.4

                        
                    if random.random() < gear_drop_prob:
                        gear_type = random.choice(["Weapons", "Armor"])
                        gear = self.generate_gear(gear_type, stage)
                        gear_name = f"{gear['TypeModifier']} {gear['ColorModifier']} {gear['BaseName']}"
                        #loot.append(gear_name) #No need to keep old weapons in inventory. We're already equipping the best stuff.
                        log.append(f"✧ Found {gear_name}")
                        gear_item = {
                            "TypeModifier": gear["TypeModifier"],
                            "ColorModifier": gear["ColorModifier"],
                            "BaseName": gear["BaseName"],
                            "type_index": gear["type_index"],
                            "basename_index": gear["basename_index"]
                        }
                        if gear_type == "Weapons" and self.adventurer.is_better_gear(gear_item, temp_state["equipped_weapon"], "Weapons", self.content):
                            temp_state["equipped_weapon"] = gear_item
                            log.append(f"✧ Equipped {gear_name}")
                        elif gear_type == "Armor" and self.adventurer.is_better_gear(gear_item, temp_state["equipped_armor"], "Armor", self.content):
                            temp_state["equipped_armor"] = gear_item
                            log.append(f"✧ Equipped {gear_name}")
                        else:
                            log.append(f"✧ Discarded {gear_name}. It's not as strong as what's already equipped.")
                    
                    elif random.random() < consumable_drop_prob:
                        #Harmful items cost less than beneficial items, so this weighting strategy doesn't work.
                        # Consumable drop (5% chance, rarity ∝ 1/Cost)
                        #total_inverse_cost = sum(1/int(item["Cost"]) for item in self.content["Items"])
                        #weights = [1/int(item["Cost"])/total_inverse_cost for item in self.content["Items"]]
                        #item = random.choices(self.content["Items"], weights=weights, k=1)[0]
                        item = random.choice(self.content["Items"])
                        log.append(f"+ Found {item['ItemName']}")
                        if random.random() <= 0.55:
                            temp_state["inventory"].append(item["ItemName"])
                            log.append(f": Kept {item['ItemName']}: {item['TargetStat']} {item['Effect']}")
                        else:
                            log.append(f": Used {item['ItemName']}: {item['TargetStat']} {'+' if item['Effect'] >= 0 else ''}{item['Effect']}")
                            consumable_effects[item["TargetStat"]] += item["Effect"]
                        current_hp += consumable_effects["HP"]
                    break
               
                #Enemy Attacks
                enemy_attack_factor = stage + 1
                attack_damage = enemy["attack"] + random.randint(-enemy_attack_factor, enemy_attack_factor)
                current_hp -= attack_damage
                log.append(f"  - {enemy['name']} deals {attack_damage} damage to {self.adventurer.name} (HP: {current_hp})")

            
            hp_changes.append(current_hp)
            
            if current_hp <= 0:
                break
        #outside while hp>0 loop
        if current_hp <= 0:
            log.append(f"☠️ {self.adventurer.name} has fallen!")
            
        temp_state["xp"] = int(temp_state["xp"])
        log.append(f"/// Run Complete")
        log.append(f"+ Defeated {killcount} enemies.")
        log.append(f"+ Gained {temp_state['xp']} XP (Total: {self.adventurer.xp + temp_state['xp']})")
        temp_state["tasks_completed"] = 1
        return log, hp_changes, temp_state

    def prune_and_prepend_log(self, log):
        try:
            with open(self.log_file, "r", encoding="utf-8") as f:
                content = f.read()
            adventures = content.strip().split("\n\n") if content.strip() else []
        except FileNotFoundError:
            adventures = []

        # Prepend the new adventure
        adventures = [ "\n".join(log) ] + adventures

        # Trim to the most recent MAX_ADVENTURES
        adventures = adventures[:MAX_ADVENTURES]

        with open(self.log_file, "w", encoding="utf-8") as f:
            f.write("\n\n".join(adventures) + "\n\n")

    def prune_log(self):
        try:
            with open(self.log_file, "r", encoding="utf-8") as f:
                content = f.read()
            adventures = content.strip().split("\n\n")
            if len(adventures) >= MAX_ADVENTURES:
                adventures = adventures[-MAX_ADVENTURES + 1:]
                with open(self.log_file, "w", encoding="utf-8") as f:
                    f.write("\n\n".join(adventures) + "\n\n")
        except FileNotFoundError:
            with open(self.log_file, "w", encoding="utf-8") as f:
                f.write("")

    def generate_gear(self, gear_type, stage):
        gear = self.content[gear_type]
        max_index = len(gear["TypeModifier"])
        difficulty_level = ((stage+self.adventurer.level)/2) // STAGES_PER_GEAR_LEVEL

        min_difficulty = min(max_index-1,max(0, difficulty_level - (DIFFICULTY_RANGE*5)))
        max_difficulty = min(max_index, difficulty_level + DIFFICULTY_RANGE)
        bias_factor = 0.3
        mode_difficulty = int(min_difficulty + (max_difficulty - min_difficulty) * bias_factor)
        
        #Basenames have a smaller range than type, because they are greater magnitude
        basename_min = min_difficulty // 3
        basename_max = max_difficulty // 3
        basename_mode = mode_difficulty // 3
        
        type_difficulty = round(random.triangular(int(min_difficulty), int(max_difficulty), mode_difficulty))
        basename_difficulty = round(random.triangular(basename_min, basename_max, basename_mode))
        type_index = min((type_difficulty) % (max_index+1), max_index - 1)
        basename_index = min((basename_difficulty) % (max_index+1), max_index - 1)
        
        type_mod = gear["TypeModifier"][type_index]
        color_mod = random.choice(gear["ColorModifier"])
        base_name = gear["BaseName"][basename_index]
        
        return {
            "TypeModifier": type_mod,
            "ColorModifier": color_mod,
            "BaseName": base_name,
            "type_index": type_index,
            "basename_index": basename_index
        }

    def show_achievement_rewards(self, achievement):
        if not self.stats_window or not self.stats_window.winfo_exists():
            return
        
        def apply_reward(effect_stat, effect):
            if effect_stat == "HP":
                self.adventurer.base_hp += int(effect)
            elif effect_stat == "Attack":
                self.adventurer.base_attack += int(effect)
            elif effect_stat == "XP":
                self.adventurer.xp += int(effect)
            messagebox.showinfo("Success", f"Received {effect} {effect_stat} for {achievement['Name']}!")            
            finalize_achievement()
        
        def refuse_reward():
            # player gets no reward for this achievement.
            finalize_achievement()
            messagebox.showinfo("Success", f"Achievement {achievement['Name']} unlocked!")
        
        def postpone_reward():
            # do nothing now, the achievement will be offered again later.
            dialog.destroy()
        
        def finalize_achievement():
            self.adventurer.achievements.append(achievement["Name"])
            self.adventurer.achievements_awarded[achievement["Name"]] += 1
            self.save_adventurer()
            dialog.destroy()

            
        dialog = Toplevel(self.stats_window)
        dialog.title(f"Achievement: {achievement['Name']}")
        dialog.geometry("400x250")
        dialog.transient(self.stats_window)
        dialog.grab_set()
        
        Label(dialog, text=f"Choose your reward for {achievement['Name']}:").pack(pady=10)
        
        #if there are effects, show buttons for them.
        reward_buttons_created = False
        for effect_stat, effect in zip(achievement["Effect Stat"], achievement["Effect"]):
            if effect_stat != "None":
                reward_buttons_created = True
                button_text = f"+{effect} {effect_stat}"
                ttk.Button(dialog, text=button_text, command=lambda es=effect_stat, e=effect: apply_reward(es, e)).pack(pady=5)
        
        #else show no reward button.
        if not reward_buttons_created:
            ttk.Button(dialog, text="Accept Achievement (No Reward)", command=skip_reward).pack(pady=5)
        
        #always allow postpone
        ttk.Button(dialog, text="Postpone", command=postpone_reward).pack(pady=5)

    def check_achievements(self, log=None):
        # called from show_adventurer_window, as soon as the window opens.
        pending_achievements = []
        award_count = {}
        for achievement in self.content["Achievements"]:
            achievement_name = achievement["Name"]
            cost_stat = achievement["Cost Stat"]
            cost = achievement["Cost"]
            repeats = achievement["Repeats"]
            self.adventurer.achievement_progress.setdefault(cost_stat, 0)
            self.adventurer.achievements_awarded.setdefault(achievement_name, 0)
            award_count.setdefault(achievement_name, self.adventurer.achievements_awarded[achievement_name])
            self.adventurer.achievements_awarded[achievement_name]
            
            if (self.adventurer.achievement_progress[cost_stat] >= cost and
                award_count[achievement_name] < repeats):
                pending_achievements.append(achievement)
                award_count[achievement_name] += 1
                if achievement["Effect Stat"] == ["None"]:
                    self.adventurer.achievements.append(achievement_name)
                    if log:
                        log.append(f"Achievement Unlocked: {achievement_name}!")
        return pending_achievements, log

    def refresh_adventurer_window(self, notebook):
        #recreate window while preserving position and currently opened tab.
        if not self.stats_window or not self.stats_window.winfo_exists():
            return

        current_tab_index = notebook.index("current")
        geometry = self.stats_window.geometry()

        self.populate_adventurer_window_contents(notebook)

        notebook.select(current_tab_index)
        self.stats_window.geometry(geometry)
        self.stats_window.lift()

    def show_adventurer_window(self):
        #create the window, define behavior, create notebook for tabs, call to populate, and then slap refresh button on top.
        if not Adventure_Feature_Enabled:
            return
        if self.stats_window and self.stats_window.winfo_exists():
            self.stats_window.lift()
            return
        self.stats_window = Toplevel(self.task_manager.root)
        self.stats_window.title("Manage Stats")
        self.stats_window.geometry("800x650")
        def on_close():
            self.stats_window.destroy()
            self.stats_window = None
        self.stats_window.protocol("WM_DELETE_WINDOW", on_close)

        #def on_refresh():
        #    on_close()
        #    self.show_adventurer_window()
        
        notebook = ttk.Notebook(self.stats_window)
        notebook.pack(fill="both", expand=True)
        
        self.populate_adventurer_window_contents(notebook)
        
        #ttk.Button(self.stats_window, text="Refresh", command=on_refresh).pack(side="right", padx=5, pady=5)
        ttk.Button(self.stats_window, text="Refresh", command=lambda:self.refresh_adventurer_window(notebook)).place(relx=1.0, rely=0.0, anchor="ne", x=-1, y=-1)
        
    def populate_adventurer_window_contents(self, notebook):
        #show all the crap for the adventurer window
        for tab in notebook.winfo_children():
            tab.destroy()
            
        # Stats Overview Tab
        stats_frame = ttk.Frame(notebook)
        notebook.add(stats_frame, text="Stats")
        
        def add_stat_label(row, label, value):
            Label(stats_frame, text=label, anchor="w").grid(row=row, column=0, sticky="w", padx=5, pady=2)
            Label(stats_frame, text=value, anchor="w").grid(row=row, column=1, sticky="w", padx=5, pady=2)
        
        row = 0
        # Name input and update button
        Label(stats_frame, text="Name:", anchor="w").grid(row=row, column=0, sticky="w", padx=5, pady=2)
        name_entry = ttk.Entry(stats_frame)
        name_entry.insert(0, self.adventurer.name)
        name_entry.grid(row=row, column=1, sticky="w", padx=5, pady=2)
        def update_name():
            new_name = name_entry.get().strip()
            if not new_name:
                messagebox.showerror("Invalid Name", "Name cannot be empty!")
                return
            # Update leaderboard first to remove old name
            self.leaderboard = [entry for entry in self.leaderboard if entry["name"] != self.adventurer.name]
            self.adventurer.name = new_name
            self.leaderboard.append({"name": new_name, "xp": self.adventurer.xp})
            self.leaderboard.sort(key=lambda x: x["xp"], reverse=True)
            self.save_adventurer()
            self.save_leaderboard()
            messagebox.showinfo("Success", f"Name changed to {new_name}!")
            on_refresh()  # Refresh UI to update labels
        ttk.Button(stats_frame, text="Update Name", command=update_name).grid(row=row, column=2, sticky="w", padx=5, pady=2)
        
        row += 1
        add_stat_label(row, "Level:", self.adventurer.level)
        row += 1
        add_stat_label(row, "XP:", self.adventurer.xp)
        row += 1
        add_stat_label(row, "XP to Next Level:", 3000 + (self.adventurer.level * 10)) #"xp_needed" is how it's usually tagged
        row += 1
        add_stat_label(row, "Tasks Completed:", self.adventurer.tasks_completed)
        row += 1
        add_stat_label(row, "Base HP:", self.adventurer.base_hp)
        row += 1
        add_stat_label(row, "Base Attack:", self.adventurer.base_attack)
        row += 1
        add_stat_label(row, "Enemies Defeated:", self.adventurer.enemy_defeats)
        row += 1
        add_stat_label(row, "Achievements:", len(self.adventurer.achievements))
        row += 1

        weapon_damage, weapon_variation = get_weapon_strength(self.adventurer.equipped_weapon, self.content)
        weapon_name = (f"{self.adventurer.equipped_weapon['TypeModifier']} {self.adventurer.equipped_weapon['ColorModifier']} {self.adventurer.equipped_weapon['BaseName']}" 
                       if self.adventurer.equipped_weapon else "None")
        weapon_display = f"{weapon_name} ({weapon_damage}±{weapon_variation*100}%)" if self.adventurer.equipped_weapon else "None"
        add_stat_label(row, "Equipped Weapon:", weapon_display)
        row += 1

        armor_HP, armor_variation = get_armor_strength(self.adventurer.equipped_armor, self.content)
        armor_name = (f"{self.adventurer.equipped_armor['TypeModifier']} {self.adventurer.equipped_armor['ColorModifier']} {self.adventurer.equipped_armor['BaseName']}" 
                      if self.adventurer.equipped_armor else "None")
        armor_display = f"{armor_name} ({armor_HP}±{armor_variation})" if self.adventurer.equipped_armor else "None"
        add_stat_label(row, "Equipped Armor:", armor_display)
        row += 1
        #old way...
        #inventory_text = ", ".join(self.adventurer.inventory) if self.adventurer.inventory else "Empty"
        #add_stat_label(row, "Inventory:", inventory_text)


        #---------------------Items from Last Adventure
        
        recent_frame = ttk.LabelFrame(stats_frame, text="Items from Last Adventure")
        recent_frame.grid(row=row, column=0, columnspan=3, sticky="ew", padx=5, pady=5)
        hp_recent = len([i for i in self.adventurer.recent_items if next((item for item in self.content["Items"] if item["ItemName"] == i), None)["TargetStat"] == "HP"])
        attack_recent = len([i for i in self.adventurer.recent_items if next((item for item in self.content["Items"] if item["ItemName"] == i), None)["TargetStat"] == "Attack"])
        Label(recent_frame, text=f"HP Items: {hp_recent}", font=("Segoe UI", 9)).pack(anchor="w", padx=5, pady=2)
        Label(recent_frame, text=f"Attack Items: {attack_recent}", font=("Segoe UI", 9)).pack(anchor="w", padx=5, pady=2)
        button_frame = ttk.Frame(recent_frame)
        button_frame.pack(fill="x", pady=2)
        if hp_recent > 0: ttk.Button(button_frame, text="Use HP Item", command=lambda: self.use_recent_item("HP")).pack(side="left", padx=2)
        if attack_recent > 0: ttk.Button(button_frame, text="Use Attack Item", command=lambda: self.use_recent_item("Attack")).pack(side="left", padx=2)
        if hp_recent > 0: ttk.Button(button_frame, text="Trash HP Item", command=lambda: self.trash_recent_item("HP")).pack(side="left", padx=2)
        if attack_recent > 0: ttk.Button(button_frame, text="Trash Attack Item", command=lambda: self.trash_recent_item("Attack")).pack(side="left", padx=2)
        row += 1
         
        #---------------------Stored Items

        stored_frame = ttk.LabelFrame(stats_frame, text="Items Stored")
        stored_frame.grid(row=row, column=0, columnspan=3, sticky="ew", padx=5, pady=5)
        hp_stored = len([i for i in self.adventurer.inventory if next((item for item in self.content["Items"] if item["ItemName"] == i), None)["TargetStat"] == "HP"])
        attack_stored = len([i for i in self.adventurer.inventory if next((item for item in self.content["Items"] if item["ItemName"] == i), None)["TargetStat"] == "Attack"])
        Label(stored_frame, text=f"HP Items: {hp_stored}", font=("Segoe UI", 9)).pack(anchor="w", padx=5, pady=2)
        Label(stored_frame, text=f"Attack Items: {attack_stored}", font=("Segoe UI", 9)).pack(anchor="w", padx=5, pady=2)
        button_frame = ttk.Frame(stored_frame)
        button_frame.pack(fill="x", pady=2)
        if hp_stored > 0: ttk.Button(button_frame, text="Use HP Item", command=lambda: self.use_stored_item("HP")).pack(side="left", padx=2)
        if attack_stored > 0: ttk.Button(button_frame, text="Use Attack Item", command=lambda: self.use_stored_item("Attack")).pack(side="left", padx=2)
        if hp_stored > 0: ttk.Button(button_frame, text="Trash HP Item", command=lambda: self.trash_stored_item("HP")).pack(side="left", padx=2)
        if attack_stored > 0: ttk.Button(button_frame, text="Trash Attack Item", command=lambda: self.trash_stored_item("Attack")).pack(side="left", padx=2)
        if hp_stored > 0: ttk.Button(button_frame, text="Save HP Item", command=lambda: self.save_stored_item("HP")).pack(side="left", padx=2)
        if attack_stored > 0: ttk.Button(button_frame, text="Save Attack Item", command=lambda: self.save_stored_item("Attack")).pack(side="left", padx=2)
        row += 1
        
        #---------------------Saved Items
        
        saved_frame = ttk.LabelFrame(stats_frame, text="Items Saved")
        saved_frame.grid(row=row, column=0, columnspan=3, sticky="ew", padx=5, pady=5)
        hp_stored = len([i for i in self.adventurer.saved_items if next((item for item in self.content["Items"] if item["ItemName"] == i), None)["TargetStat"] == "HP"])
        attack_stored = len([i for i in self.adventurer.saved_items if next((item for item in self.content["Items"] if item["ItemName"] == i), None)["TargetStat"] == "Attack"])
        Label(saved_frame, text=f"HP Items: {hp_stored}", font=("Segoe UI", 9)).pack(anchor="w", padx=5, pady=2)
        Label(saved_frame, text=f"Attack Items: {attack_stored}", font=("Segoe UI", 9)).pack(anchor="w", padx=5, pady=2)
        button_frame = ttk.Frame(saved_frame)
        button_frame.pack(fill="x", pady=2)
        if hp_stored > 0: ttk.Button(button_frame, text="Use HP Item", command=lambda: self.use_saved_item("HP")).pack(side="left", padx=2)
        if attack_stored > 0: ttk.Button(button_frame, text="Use Attack Item", command=lambda: self.use_saved_item("Attack")).pack(side="left", padx=2)
        if hp_stored > 0: ttk.Button(button_frame, text="Trash HP Item", command=lambda: self.trash_saved_item("HP")).pack(side="left", padx=2)
        if attack_stored > 0: ttk.Button(button_frame, text="Trash Attack Item", command=lambda: self.trash_saved_item("Attack")).pack(side="left", padx=2)
        row += 1
        
        xp_label = Label(stats_frame, text=f"XP: {self.adventurer.xp}", anchor="w")
        xp_label.grid(row=2, column=1, sticky="w", padx=5, pady=2)
        #inventory_label = Label(stats_frame, text=inventory_text, anchor="w")
        #inventory_label.grid(row=11, column=1, sticky="w", padx=5, pady=2)


        # Trigger Dialog for pending achievements
        pending_achievements, log = self.check_achievements(log=None)
        for achievement in pending_achievements:
            if achievement["Effect Stat"] != ["None"]:
                self.task_manager.root.after(100, lambda: self.show_achievement_rewards(achievement))

        
        # Achievements Tab
        achievements_frame = ttk.Frame(notebook)
        notebook.add(achievements_frame, text="Achievements")
        
        achievements_tree = ttk.Treeview(achievements_frame, columns=("Name", "Count"), show="headings")
        achievements_tree.heading("Name", text="Name")
        achievements_tree.heading("Count", text="Count")
        achievements_tree.pack(fill="both", expand=True)
        
        achievement_counts = {}
        for ach in self.adventurer.achievements:
            achievement_counts[ach] = achievement_counts.get(ach, 0) + 1
        for name, count in achievement_counts.items():
            achievements_tree.insert("", "end", values=(name, count))
        
        # Skill Tree Tab
        skill_frame = ttk.Frame(notebook)
        notebook.add(skill_frame, text="Skills")
        
        skill_tree = ttk.Treeview(skill_frame, columns=("Name", "Cost", "Effect"), show="headings")
        skill_tree.heading("Name", text="Name")
        skill_tree.heading("Cost", text="Cost")
        skill_tree.heading("Effect", text="Effect")
        skill_tree.pack(fill="both", expand=True)
        
        for node in self.content["SkillTree"]:
            cost = f"{node['Cost']} XP"
            if node.get("AchievementReq"):
                cost += f", {node['AchievementReq']}"
            skill_tree.insert("", "end", values=(node["Name"], cost, node["Effect"]), tags=(node["ID"],))
        
        def upgrade_skill():
            selected = skill_tree.selection()
            if not selected:
                messagebox.showwarning("No Selection", "Please select a skill to upgrade.")
                return
            node_id = skill_tree.item(selected[0], "tags")[0]
            node = next(n for n in self.content["SkillTree"] if n["ID"] == node_id)
            
            if node.get("AchievementReq") and node["AchievementReq"] not in self.adventurer.achievements:
                messagebox.showerror("Locked", f"Requires achievement: {node['AchievementReq']}")
                return
            if self.adventurer.xp < node["Cost"]:
                messagebox.showerror("Insufficient XP", f"Need {node['Cost']} XP")
                return
            
            self.adventurer.xp -= node["Cost"]
            if node["Type"] == "Stat":
                if "HP" in node["Effect"]:
                    self.adventurer.base_hp += int(node["Effect"].split("+")[1])
                    hp_label.config(text=f"{self.adventurer.base_hp}")
                elif "Attack" in node["Effect"]:
                    self.adventurer.base_attack += int(node["Effect"].split("+")[1])
                    attack_label.config(text=f"{self.adventurer.base_attack}")
            elif node["Type"] == "Ability":
                self.adventurer.skills.append(node["Name"])
            elif node["Type"] == "ItemBoost":
                pass
                
            self.save_adventurer()
            xp_label.config(text=f"{self.adventurer.xp}")
            messagebox.showinfo("Success", f"Unlocked {node['Name']}!")
        
        ttk.Button(skill_frame, text="Upgrade", command=upgrade_skill).pack(pady=5)
        
        # Shop Tab
        shop_frame = ttk.Frame(notebook)
        notebook.add(shop_frame, text="Shop")
        
        sort_column = [None]  # List to allow modification in nested functions
        sort_reverse = [False]
        
        shop_tree = ttk.Treeview(shop_frame, columns=("ItemName", "TargetStat", "Effect", "Cost"), show="headings")
        
        def sort_by_column(column):
            """Sort Treeview by the specified column, toggling ascending/descending."""
            if sort_column[0] == column:
                sort_reverse[0] = not sort_reverse[0]
            else:
                sort_column[0] = column
                sort_reverse[0] = False
            populate_shop_tree()
        
        def populate_shop_tree():
            """Populate or repopulate the shop Treeview with sorted items."""
            for item in shop_tree.get_children():
                shop_tree.delete(item)
            items = self.content["Items"]
            if sort_column[0]:
                is_numeric = sort_column[0] in ["Effect", "Cost"]
                items = sorted(
                    items,
                    key=lambda x: (int(x[sort_column[0]]) if is_numeric and x[sort_column[0]] else x[sort_column[0]]) or "",
                    reverse=sort_reverse[0]
                )
            for item in items:
                shop_tree.insert("", "end", values=(item["ItemName"], item["TargetStat"], item["Effect"], item["Cost"]))
        
        shop_tree.heading("ItemName", text="Item Name", command=lambda: sort_by_column("ItemName"))
        shop_tree.heading("TargetStat", text="Target Stat", command=lambda: sort_by_column("TargetStat"))
        shop_tree.heading("Effect", text="Effect", command=lambda: sort_by_column("Effect"))
        shop_tree.heading("Cost", text="Cost (XP)", command=lambda: sort_by_column("Cost"))
        shop_tree.pack(fill="both", expand=True)
        
        populate_shop_tree()
        
        def buy_item():
            selected = shop_tree.selection()
            if not selected:
                messagebox.showwarning("No Selection", "Please select an item to buy.")
                return
            item_name = shop_tree.item(selected[0], "values")[0]
            item = next(i for i in self.content["Items"] if i["ItemName"] == item_name)
            cost = int(item["Cost"])
            if self.adventurer.xp < cost:
                messagebox.showerror("Insufficient XP", f"Need {cost} XP to buy {item_name}")
                return
            self.adventurer.xp -= cost
            self.adventurer.inventory.append(item_name)
            self.save_adventurer()
            xp_label.config(text=f"XP: {self.adventurer.xp}")
            messagebox.showinfo("Success", f"Bought {item_name}! It has been placed in your 'Items Stored'.")
            self.refresh_adventurer_window(self.stats_window.winfo_children()[0] if self.stats_window else None)

        ttk.Button(shop_frame, text="Buy", command=buy_item).pack(pady=5)
        
        # Adventure Log Tab
        log_frame = ttk.Frame(notebook)
        notebook.add(log_frame, text="Adventure Log")
        
        log_text = Text(log_frame, height=20, wrap="word")
        log_text.pack(side="left", fill="both", expand=True)
        scrollbar = Scrollbar(log_frame, command=log_text.yview)
        scrollbar.pack(side="right", fill="y")
        log_text.config(yscrollcommand=scrollbar.set)
        
        try:
            with open(self.log_file, "r", encoding="utf-8") as f:
                log_text.insert("1.0", f.read())
        except FileNotFoundError:
            log_text.insert("1.0", "No adventures yet.")
            with open(self.log_file, "w", encoding="utf-8") as f:
                f.write("")
        log_text.config(state="disabled")
        
                # Narrative History Tab
        narrative_frame = ttk.Frame(notebook)
        notebook.add(narrative_frame, text="Narrative History")
        
        narrative_text = Text(narrative_frame, height=20, wrap="word")
        narrative_text.pack(side="left", fill="both", expand=True)
        narrative_scrollbar = Scrollbar(narrative_frame, command=narrative_text.yview)
        narrative_scrollbar.pack(side="right", fill="y")
        narrative_text.config(yscrollcommand=narrative_scrollbar.set)
 
#Grok made the following flourish.... maybe I'll use it? Depends how the other one ends up looking in practice. 
#╔═══════╦═══════════════════════════════════════════════════╦═══════╗
#║*~♦~*  ║                                                   ║  *~♦~*║
#╠═══════╩═══════════════════════════════════════════════════╩═══════╣
#║*~♦~*                                                         *~♦~*║
#╠═══════╦═══════════════════════════════════════════════════╦═══════╣
#║*~♦~*  ║                                                   ║  *~♦~*║
#╚═══════╩═══════════════════════════════════════════════════╩═══════╝"
        
        
        # ASCII flourish from get_narrative_event
        flourish_top =      "╭══════════════════════════════ ༺ ✦ ༻ ══════════════════════════════╮\n║♦ STORY EVENT"
        flourish_body =     "\n║♦ "
        flourish_bottom =   "\n╰══════════════════════════════ ༺ ✦ ༻ ══════════════════════════════╯\n"
        
        # Display all encountered narrative events
        if self.content.get("SequentialNarrative") and self.adventurer.narrative_progress > 0:
            for i in range(self.adventurer.narrative_progress):
                try:
                    event = self.content["SequentialNarrative"][i]["Description"]
                    wrapped_event = textwrap.fill(event, width=60)
                    centered_event = [line.center(60) for line in wrapped_event.split('\n')]
                    formatted_event = f"{flourish_top}{flourish_body}{flourish_body.join(centered_event)}{flourish_bottom}"
                    narrative_text.insert("end", formatted_event)
                except Exception as E:
                    print("Failed to run narrative Event:", E)
        else:
            narrative_text.insert("end", "No narrative events encountered yet.\n")
        
        narrative_text.config(state="disabled")
        
        # Leaderboard Tab
        leaderboard_frame = ttk.Frame(notebook)
        notebook.add(leaderboard_frame, text="Leaderboard")
        
        lb_tree = ttk.Treeview(leaderboard_frame, columns=("Rank", "Name", "XP"), show="headings")
        lb_tree.heading("Rank", text="Rank")
        lb_tree.heading("Name", text="Name")
        lb_tree.heading("XP", text="XP")
        lb_tree.pack(fill="both", expand=True)
        
        for i, entry in enumerate(self.leaderboard[:10], 1):
            lb_tree.insert("", "end", values=(i, entry["name"], entry["xp"]))
                # Add refresh button in top-right corner


